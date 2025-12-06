# -*- coding: utf-8 -*-
"""Export FX Graph to Excel"""

import torch
import torch.fx as fx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Optional, Dict, Any, List


def export_excel(gm: fx.GraphModule, path: str) -> None:
    """Export FX GraphModule to Excel file"""
    wb = Workbook()

    # Sheet 1: Overview
    _create_overview_sheet(wb, gm)

    # Sheet 2: Nodes
    _create_nodes_sheet(wb, gm)

    # Sheet 3: Parameters
    _create_params_sheet(wb, gm)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)


def _create_overview_sheet(wb: Workbook, gm: fx.GraphModule) -> None:
    """Create overview sheet"""
    ws = wb.create_sheet("Overview")

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Count nodes by type
    node_counts: Dict[str, int] = {}
    input_count = 0
    output_count = 0

    for node in gm.graph.nodes:
        if node.op == "placeholder":
            input_count += 1
        elif node.op == "output":
            output_count += 1
        elif node.op == "call_function":
            op_name = _get_op_name(node.target)
            node_counts[op_name] = node_counts.get(op_name, 0) + 1
        elif node.op == "call_module":
            node_counts[node.target] = node_counts.get(node.target, 0) + 1

    # Count parameters
    param_count = 0
    param_size = 0
    for name, param in gm.named_buffers():
        param_count += 1
        param_size += param.numel() * param.element_size()

    # Write overview
    data = [
        ["Model Overview", ""],
        ["", ""],
        ["Total Nodes", len(list(gm.graph.nodes))],
        ["Inputs", input_count],
        ["Outputs", output_count],
        ["Parameters", param_count],
        ["Parameter Size (MB)", f"{param_size / 1024 / 1024:.2f}"],
        ["", ""],
        ["Operator Statistics", "Count"],
    ]

    for i, row in enumerate(data, 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
            if i == 1 or i == 9:
                ws.cell(row=i, column=j).fill = header_fill
                ws.cell(row=i, column=j).font = header_font

    # Write op counts
    row = 10
    for op_name, count in sorted(node_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=op_name)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Adjust column width
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


def _create_nodes_sheet(wb: Workbook, gm: fx.GraphModule) -> None:
    """Create nodes sheet"""
    ws = wb.create_sheet("Nodes")

    # Header
    headers = ["#", "Name", "Op Type", "Target", "Inputs", "Output Shape", "Attributes"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for j, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write nodes
    row = 2
    for i, node in enumerate(gm.graph.nodes):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=node.name)
        ws.cell(row=row, column=3, value=node.op)
        ws.cell(row=row, column=4, value=_get_target_str(node.target))
        ws.cell(row=row, column=5, value=_get_inputs_str(node))
        ws.cell(row=row, column=6, value=_get_shape_str(node))
        ws.cell(row=row, column=7, value=_get_attrs_str(node))
        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 30


def _create_params_sheet(wb: Workbook, gm: fx.GraphModule) -> None:
    """Create parameters sheet"""
    ws = wb.create_sheet("Parameters")

    # Header
    headers = ["#", "Name", "Shape", "Dtype", "Elements", "Size (KB)", "Min", "Max", "Mean", "Std"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for j, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write parameters
    row = 2
    for i, (name, param) in enumerate(gm.named_buffers()):
        size_kb = param.numel() * param.element_size() / 1024
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=str(list(param.shape)))
        ws.cell(row=row, column=4, value=str(param.dtype))
        ws.cell(row=row, column=5, value=param.numel())
        ws.cell(row=row, column=6, value=f"{size_kb:.2f}")

        if param.dtype in [torch.float32, torch.float16, torch.float64]:
            ws.cell(row=row, column=7, value=f"{param.min().item():.4f}")
            ws.cell(row=row, column=8, value=f"{param.max().item():.4f}")
            ws.cell(row=row, column=9, value=f"{param.mean().item():.4f}")
            ws.cell(row=row, column=10, value=f"{param.std().item():.4f}")

        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12


def _get_op_name(target) -> str:
    """Get readable operator name"""
    if callable(target):
        if hasattr(target, "__name__"):
            return target.__name__
        elif hasattr(target, "__class__"):
            return target.__class__.__name__
    return str(target)


def _get_target_str(target) -> str:
    """Get target as string"""
    if callable(target):
        if hasattr(target, "__module__") and hasattr(target, "__name__"):
            return f"{target.__module__}.{target.__name__}"
        elif hasattr(target, "__name__"):
            return target.__name__
    return str(target)


def _get_inputs_str(node: fx.Node) -> str:
    """Get inputs as string"""
    inputs = []
    for arg in node.args:
        if isinstance(arg, fx.Node):
            inputs.append(arg.name)
        else:
            inputs.append(str(arg))
    return ", ".join(inputs)


def _get_shape_str(node: fx.Node) -> str:
    """Get output shape from meta if available"""
    if "tensor_meta" in node.meta:
        meta = node.meta["tensor_meta"]
        if hasattr(meta, "shape"):
            return str(list(meta.shape))
    if "shape" in node.meta:
        return str(node.meta["shape"])
    return ""


def _get_attrs_str(node: fx.Node) -> str:
    """Get kwargs as string"""
    if node.kwargs:
        return str(dict(node.kwargs))
    return ""
